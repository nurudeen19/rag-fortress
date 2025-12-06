"""
Document Loader - Loads documents from FileUpload model with metadata enrichment.
"""

from typing import List, Optional, Dict, Any
from pathlib import Path
import json
import csv
import os
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select

from app.config.settings import settings
from app.core import get_logger
from app.models.file_upload import FileUpload, FileStatus, SecurityLevel
from app.models.department import Department


logger = get_logger(__name__)


class DocumentLoader:
    """Loads documents from FileUpload model with enriched metadata."""
    
    def __init__(self, session: AsyncSession):
        """Initialize loader with database session."""
        self.session = session
    
    async def load_pending_files(self, file_ids: Optional[List[int]] = None) -> List[Dict[str, Any]]:
        """
        Load pending approved files from database with enriched metadata.
        
        Args:
            file_ids: Optional list of specific file IDs to load. If provided:
                     - If contains IDs: Load only those specific files
                     - If None: Load all approved files not yet processed
        
        Returns:
            List of documents with content and metadata
        """
        logger.info(f"Loading pending approved files from database (file_ids={file_ids})")
        
        # Build query based on file_ids parameter
        if file_ids and len(file_ids) > 0:
            # Load specific files by ID
            stmt = select(FileUpload).where(FileUpload.id.in_(file_ids)).order_by(FileUpload.created_at)
            logger.info(f"Loading {len(file_ids)} specific files: {file_ids}")
        else:
            # Load all pending approved files (not yet processed)
            stmt = select(FileUpload).where(
                (FileUpload.status == FileStatus.APPROVED) &
                (FileUpload.is_processed.is_(False))
            ).order_by(FileUpload.created_at)
            logger.info("Loading all pending approved files")
        
        result = await self.session.execute(stmt)
        file_uploads = result.scalars().all()
        
        logger.info(f"Found {len(file_uploads)} files to load")
        
        documents = []
        for file_upload in file_uploads:
            try:
                # Load and enrich document
                doc = await self._load_and_enrich(file_upload)
                documents.append(doc)
                logger.info(f"✓ Loaded: {file_upload.file_name}")
            except Exception as e:
                logger.error(f"✗ Failed to load {file_upload.file_name}: {e}")
        
        logger.info(f"Successfully loaded {len(documents)} documents")
        return documents
    
    async def _load_and_enrich(self, file_upload: FileUpload) -> Dict[str, Any]:
        """Load file and enrich with metadata from model."""
        # Resolve relative path to full path
        rel_path = file_upload.file_path
        if not os.path.isabs(rel_path):
            base_dir = os.getenv("FILES_DIR", "data/files")
            full_path = os.path.join(base_dir, rel_path)
        else:
            full_path = rel_path
        
        file_path = Path(full_path)
        
        # Load file content
        content = self._load_file_content(file_path, file_upload.file_type)
        
        # Get department name if applicable
        dept_name = None
        if file_upload.department_id:
            stmt = select(Department).where(Department.id == file_upload.department_id)
            result = await self.session.execute(stmt)
            dept = result.scalar_one_or_none()
            dept_name = dept.name if dept else None
        
        # Build enriched metadata
        meta = {
            "upload_token": file_upload.upload_token,
            "file_id": file_upload.id,
            # store integer representation for simpler ACL comparisons
            "security_level": file_upload.security_level.value,
            "is_department_only": file_upload.is_department_only,
            "department": dept_name,
            "file_type": file_upload.file_type,
            "department_id": file_upload.department_id or None,
            "file_purpose": file_upload.file_purpose,
            "file_size_bytes": file_upload.file_size,
        }
        
        # Parse field_selection JSON string to list
        field_selection_parsed = None
        if file_upload.field_selection:
            try:
                field_selection_parsed = json.loads(file_upload.field_selection)
            except json.JSONDecodeError:
                logger.warning(f"Failed to parse field_selection for file {file_upload.id}: {file_upload.field_selection}")
        
        return {
            "file_id": file_upload.id,
            "file_name": file_upload.file_name,
            "file_type": file_upload.file_type,
            "content": content,            
            "field_selection": field_selection_parsed,
            "meta": meta,
        }
    
    def _load_file_content(self, file_path: Path, file_type: str) -> Any:
        """Load file content based on type."""
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        file_type_lower = file_type.lower()
        
        # Text-based files
        if file_type_lower in ['txt', 'md', 'markdown']:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        
        # JSON files
        elif file_type_lower == 'json':
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        
        # CSV files
        elif file_type_lower == 'csv':
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                return list(reader)
        
        # PDF files
        elif file_type_lower == 'pdf':
            try:
                from langchain_community.document_loaders import PyPDFLoader
                loader = PyPDFLoader(str(file_path))
                docs = loader.load()
                return "\n\n".join([d.page_content for d in docs])
            except ImportError as e:
                logger.error(f"PDF loader not available. Install with: pip install pypdf")
                raise
            except Exception as e:
                logger.error(f"Failed to load PDF: {e}")
                raise
        
        # Word documents
        elif file_type_lower in ['doc', 'docx']:
            try:
                # Try python-docx first (lightweight, requires python-docx package)
                import docx
                doc = docx.Document(str(file_path))
                paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
                return "\n\n".join(paragraphs)
            except ImportError:
                # Fallback to UnstructuredWordDocumentLoader if python-docx not available
                try:
                    from langchain_community.document_loaders import UnstructuredWordDocumentLoader
                    loader = UnstructuredWordDocumentLoader(str(file_path))
                    docs = loader.load()
                    return "\n\n".join([d.page_content for d in docs])
                except ImportError:
                    logger.error(f"DOCX loader not available. Install with: pip install python-docx")
                    raise
            except Exception as e:
                logger.error(f"Failed to load DOCX: {e}")
                raise
        
        # Excel files
        elif file_type_lower in ['xlsx', 'xls']:
            try:
                # Try openpyxl first (lightweight, requires openpyxl package)
                import openpyxl
                wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
                content_parts = []
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    content_parts.append(f"Sheet: {sheet_name}\n")
                    for row in sheet.iter_rows(values_only=True):
                        row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                        if row_text.strip():
                            content_parts.append(row_text)
                wb.close()
                return "\n".join(content_parts)
            except ImportError:
                # Fallback to UnstructuredExcelLoader
                try:
                    from langchain_community.document_loaders import UnstructuredExcelLoader
                    loader = UnstructuredExcelLoader(str(file_path))
                    docs = loader.load()
                    return "\n\n".join([d.page_content for d in docs])
                except ImportError:
                    logger.error(f"Excel loader not available. Install with: pip install openpyxl")
                    raise
            except Exception as e:
                logger.error(f"Failed to load Excel: {e}")
                raise
        
        else:
            raise ValueError(f"Unsupported file type: {file_type_lower}")

